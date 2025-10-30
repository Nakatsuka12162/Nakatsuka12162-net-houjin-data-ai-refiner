from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.views.decorators.csrf import csrf_exempt
from .models import Company, Executive, Office, ResearchHistory, ExecutionHistory
import csv
import threading
import json
from django.utils import timezone

def admin_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user and user.is_staff:
            login(request, user)
            return redirect('admin_dashboard')
        messages.error(request, 'Invalid credentials')
    return render(request, 'admin/login.html')

def admin_logout(request):
    logout(request)
    return redirect('admin_login')

@login_required
def admin_dashboard(request):
    stats = {
        'total_companies': Company.objects.count(),
        'total_executives': Executive.objects.count(),
        'total_offices': Office.objects.count(),
        'recent_changes': ResearchHistory.objects.order_by('-timestamp')[:10],
        'industry_breakdown': Company.objects.values('industry').annotate(count=Count('id')).order_by('-count')[:5],
        'recent_companies': Company.objects.order_by('-created_at')[:5],
    }
    return render(request, 'admin/dashboard.html', stats)

@login_required
def company_list(request):
    search = request.GET.get('search', '')
    industry = request.GET.get('industry', '')
    
    companies = Company.objects.all()
    if search:
        companies = companies.filter(
            Q(company_name__icontains=search) |
            Q(corporate_number__icontains=search) |
            Q(representative_name__icontains=search)
        )
    if industry:
        companies = companies.filter(industry=industry)
    
    paginator = Paginator(companies.order_by('-updated_at'), 25)
    page = paginator.get_page(request.GET.get('page'))
    
    industries = Company.objects.values_list('industry', flat=True).distinct()
    
    return render(request, 'admin/company_list.html', {
        'companies': page,
        'search': search,
        'industry': industry,
        'industries': industries,
    })

@login_required
def company_detail(request, pk):
    company = get_object_or_404(Company, pk=pk)
    executives = company.executives.all().order_by('order')
    offices = company.offices.all().order_by('order')
    history = ResearchHistory.objects.filter(corporate_number=company.corporate_number).order_by('-timestamp')[:20]
    
    if request.method == 'POST':
        # Update company
        for field in ['company_name', 'representative_name', 'industry', 'address', 'phone', 'revenue', 'employee_count']:
            if field in request.POST:
                setattr(company, field, request.POST[field])
        company.save()
        messages.success(request, 'Company updated successfully')
        return redirect('company_detail', pk=pk)
    
    return render(request, 'admin/company_detail.html', {
        'company': company,
        'executives': executives,
        'offices': offices,
        'history': history,
    })

@login_required
def delete_company(request, pk):
    """Delete company from database and Google Sheets"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Only POST method allowed'}, status=405)
    
    try:
        company = get_object_or_404(Company, pk=pk)
        corporate_number = company.corporate_number
        company_name = company.company_name
        
        # Delete from Google Sheets first
        try:
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            import os
            
            # Google Sheets authentication
            scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
            creds_path = os.path.join(os.path.dirname(__file__), '..', 'credentials.json')
            
            if os.path.exists(creds_path):
                creds = ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)
                client = gspread.authorize(creds)
                
                # Open the spreadsheet
                spreadsheet_id = os.getenv('SPREADSHEET_ID', '1vfBx_vbWQHLsFaI4HTdj5iB0LH8jW4qTlFOdHacxjug')
                spreadsheet = client.open_by_key(spreadsheet_id)
                worksheet = spreadsheet.worksheet('会社リスト')
                
                # Find and delete the row with this corporate number
                try:
                    cell = worksheet.find(corporate_number)
                    if cell:
                        worksheet.delete_rows(cell.row)
                except Exception as sheet_error:
                    print(f"Error deleting from Google Sheets: {sheet_error}")
        except Exception as e:
            print(f"Google Sheets deletion failed: {e}")
            # Continue with database deletion even if Sheets deletion fails
        
        # Delete from database (will cascade delete executives and offices)
        company.delete()
        
        return JsonResponse({
            'status': 'success',
            'message': f'Successfully deleted {company_name} (Corporate Number: {corporate_number}) from database and Google Sheets'
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Error deleting company: {str(e)}'
        }, status=500)

@login_required
def executive_list(request):
    search = request.GET.get('search', '')
    executives = Executive.objects.select_related('company').all()
    
    if search:
        executives = executives.filter(
            Q(name__icontains=search) |
            Q(position__icontains=search) |
            Q(company__company_name__icontains=search)
        )
    
    paginator = Paginator(executives.order_by('company__company_name'), 50)
    page = paginator.get_page(request.GET.get('page'))
    
    return render(request, 'admin/executive_list.html', {
        'executives': page,
        'search': search,
    })

@login_required
def office_list(request):
    search = request.GET.get('search', '')
    offices = Office.objects.select_related('company').all()
    
    if search:
        offices = offices.filter(
            Q(name__icontains=search) |
            Q(address__icontains=search) |
            Q(company__company_name__icontains=search)
        )
    
    paginator = Paginator(offices.order_by('company__company_name'), 50)
    page = paginator.get_page(request.GET.get('page'))
    
    return render(request, 'admin/office_list.html', {
        'offices': page,
        'search': search,
    })

@login_required
def history_list(request):
    search = request.GET.get('search', '')
    history = ResearchHistory.objects.all()
    
    if search:
        history = history.filter(
            Q(corporate_number__icontains=search) |
            Q(changed_field__icontains=search)
        )
    
    paginator = Paginator(history.order_by('-timestamp'), 100)
    page = paginator.get_page(request.GET.get('page'))
    
    return render(request, 'admin/history_list.html', {
        'history': page,
        'search': search,
    })

@login_required
def export_companies_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="companies.csv"'
    
    writer = csv.writer(response)
    # Header row matching the required format: 企業法人番号, 正式企業名, 所在地, URL
    writer.writerow(['企業法人番号', '正式企業名', '所在地', 'URL'])
    
    for company in Company.objects.all().order_by('id'):
        # Use company_overview_url if available, otherwise fallback to url
        url = company.company_overview_url or company.url or ''
        
        writer.writerow([
            company.corporate_number or '',
            company.company_name or '',
            company.address or '',
            url
        ])
    
    return response

@login_required
def export_companies_excel(request):
    """Simple Excel export"""
    try:
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Companies"
        
        # Headers matching CSV format: 企業法人番号, 正式企業名, 所在地, URL
        headers = ['企業法人番号', '正式企業名', '所在地', 'URL']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row, company in enumerate(Company.objects.all().order_by('id'), 2):
            # Use company_overview_url if available, otherwise fallback to url
            url = company.company_overview_url or company.url or ''
            
            # Format corporate number as text to prevent scientific notation
            corp_cell = ws.cell(row=row, column=1, value=company.corporate_number or '')
            corp_cell.number_format = '@'  # Text format
            
            ws.cell(row=row, column=2, value=company.company_name or '')
            ws.cell(row=row, column=3, value=company.address or '')
            ws.cell(row=row, column=4, value=url)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="companies.xlsx"'
        return response
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="companies.xlsx"'
        return response
        
    except ImportError:
        return export_companies_csv(request)

@login_required
def export_companies_detailed_csv(request):
    """Export detailed company information in Excel format with one sheet per company"""
    try:
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        companies = Company.objects.all().order_by('id')
        
        for company in companies:
            # Get related executives and offices
            executives = company.executives.all().order_by('order')
            offices = company.offices.all().order_by('order')
            
            # Create sheet name from company name (max 31 chars, no invalid chars)
            sheet_name = (company.company_name or f'Company {company.id}')[:31]
            sheet_name = sheet_name.replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
            ws = wb.create_sheet(title=sheet_name)
            
            row = 1
            
            # ◆ I. 基本法人情報（識別・概要）
            ws.cell(row=row, column=1, value='◆ I. 基本法人情報（識別・概要）')
            ws.cell(row=row, column=2, value='法人番号')
            ws.cell(row=row, column=3, value=company.corporate_number or '')
            row += 1
            
            ws.cell(row=row, column=2, value='会社名')
            ws.cell(row=row, column=3, value=company.company_name or '')
            row += 1
            
            ws.cell(row=row, column=2, value='会社名かな')
            ws.cell(row=row, column=3, value=company.company_name_kana or '')
            row += 1
            
            ws.cell(row=row, column=2, value='英文企業名')
            ws.cell(row=row, column=3, value=company.english_name or '')
            row += 1
            
            ws.cell(row=row, column=2, value='代表者名')
            ws.cell(row=row, column=3, value=company.representative_name or '')
            row += 1
            
            ws.cell(row=row, column=2, value='代表者かな')
            ws.cell(row=row, column=3, value=company.representative_kana or '')
            row += 1
            
            ws.cell(row=row, column=2, value='代表者年齢')
            ws.cell(row=row, column=3, value=company.representative_age or '')
            row += 1
            
            ws.cell(row=row, column=2, value='代表者生年月日')
            ws.cell(row=row, column=3, value=company.representative_birth or '')
            row += 1
            
            ws.cell(row=row, column=2, value='代表者出身大学')
            ws.cell(row=row, column=3, value=company.representative_university or '')
            row += 1
            
            ws.cell(row=row, column=2, value='郵便番号')
            ws.cell(row=row, column=3, value=company.postal_code or '')
            row += 1
            
            ws.cell(row=row, column=2, value='住所')
            ws.cell(row=row, column=3, value=company.address or '')
            row += 1
            
            ws.cell(row=row, column=2, value='電話番号')
            ws.cell(row=row, column=3, value=company.phone or '')
            row += 1
            
            ws.cell(row=row, column=2, value='登記住所')
            ws.cell(row=row, column=3, value=company.registered_address or '')
            row += 1
            
            ws.cell(row=row, column=2, value='FAX番号')
            ws.cell(row=row, column=3, value=company.fax or '')
            row += 1
            
            ws.cell(row=row, column=2, value='URL')
            ws.cell(row=row, column=3, value=company.url or '')
            row += 1
            
            ws.cell(row=row, column=2, value='創業')
            ws.cell(row=row, column=3, value=company.founded or '')
            row += 1
            
            ws.cell(row=row, column=2, value='設立')
            ws.cell(row=row, column=3, value=company.established or '')
            row += 1
            
            ws.cell(row=row, column=2, value='資本金')
            ws.cell(row=row, column=3, value=company.capital or '')
            row += 1
            
            ws.cell(row=row, column=2, value='出資金')
            ws.cell(row=row, column=3, value=company.investment or '')
            row += 1
            
            ws.cell(row=row, column=2, value='会員数')
            ws.cell(row=row, column=3, value=company.member_count or '')
            row += 1
            
            ws.cell(row=row, column=2, value='組合員数')
            ws.cell(row=row, column=3, value=company.union_member_count or '')
            row += 1
            
            ws.cell(row=row, column=2, value='上場市場')
            ws.cell(row=row, column=3, value=company.stock_market or '')
            row += 1
            
            ws.cell(row=row, column=2, value='証券コード')
            ws.cell(row=row, column=3, value=company.stock_code or '')
            row += 1
            
            ws.cell(row=row, column=2, value='決算期')
            ws.cell(row=row, column=3, value=company.fiscal_year_end or '')
            row += 1
            
            # ◆ II. 経営・財務情報
            ws.cell(row=row, column=1, value='◆ II. 経営・財務情報')
            ws.cell(row=row, column=2, value='売上高')
            ws.cell(row=row, column=3, value=company.revenue or '')
            row += 1
            
            ws.cell(row=row, column=2, value='純利益')
            ws.cell(row=row, column=3, value=company.net_profit or '')
            row += 1
            
            ws.cell(row=row, column=2, value='預金量')
            ws.cell(row=row, column=3, value=company.deposits or '')
            row += 1
            
            ws.cell(row=row, column=2, value='従業員数')
            ws.cell(row=row, column=3, value=company.employee_count or '')
            row += 1
            
            ws.cell(row=row, column=2, value='平均年齢')
            ws.cell(row=row, column=3, value=company.average_age or '')
            row += 1
            
            ws.cell(row=row, column=2, value='平均年収')
            ws.cell(row=row, column=3, value=company.average_salary or '')
            row += 1
            
            ws.cell(row=row, column=2, value='役員数')
            ws.cell(row=row, column=3, value=company.executive_count or '')
            row += 1
            
            ws.cell(row=row, column=2, value='株主数')
            ws.cell(row=row, column=3, value=company.shareholder_count or '')
            row += 1
            
            ws.cell(row=row, column=2, value='取引銀行')
            ws.cell(row=row, column=3, value=company.main_bank or '')
            row += 1
            
            # ◆ III. 事業・業務内容
            ws.cell(row=row, column=1, value='◆ III. 事業・業務内容')
            ws.cell(row=row, column=2, value='業種')
            ws.cell(row=row, column=3, value=company.industry or '')
            row += 1
            
            ws.cell(row=row, column=2, value='事業内容')
            ws.cell(row=row, column=3, value=company.business_content or '')
            row += 1
            
            ws.cell(row=row, column=2, value='主要事業')
            ws.cell(row=row, column=3, value=company.main_business or '')
            row += 1
            
            ws.cell(row=row, column=2, value='事業エリア')
            ws.cell(row=row, column=3, value=company.business_area or '')
            row += 1
            
            ws.cell(row=row, column=2, value='系列')
            ws.cell(row=row, column=3, value=company.group_affiliation or '')
            row += 1
            
            ws.cell(row=row, column=2, value='販売先')
            ws.cell(row=row, column=3, value=company.sales_destination or '')
            row += 1
            
            ws.cell(row=row, column=2, value='仕入先')
            ws.cell(row=row, column=3, value=company.supplier or '')
            row += 1
            
            # ◆ IV. 役員名簿
            ws.cell(row=row, column=1, value='◆ IV. 役員名簿')
            ws.cell(row=row, column=2, value='役職名1')
            ws.cell(row=row, column=3, value=executives[0].position if len(executives) > 0 else '')
            row += 1
            
            ws.cell(row=row, column=2, value='役員名1')
            ws.cell(row=row, column=3, value=executives[0].name if len(executives) > 0 else '')
            row += 1
            
            ws.cell(row=row, column=2, value='ふりがな1')
            ws.cell(row=row, column=3, value=executives[0].name_kana if len(executives) > 0 else '')
            row += 1
            
            for i in range(1, 15):  # 2 to 15
                if i < len(executives):
                    ws.cell(row=row, column=2, value=f'役職名{i+1}')
                    ws.cell(row=row, column=3, value=executives[i].position or '')
                    row += 1
                    ws.cell(row=row, column=2, value=f'役員名{i+1}')
                    ws.cell(row=row, column=3, value=executives[i].name or '')
                    row += 1
                    ws.cell(row=row, column=2, value=f'ふりがな{i+1}')
                    ws.cell(row=row, column=3, value=executives[i].name_kana or '')
                    row += 1
                else:
                    ws.cell(row=row, column=2, value=f'役職名{i+1}')
                    row += 1
                    ws.cell(row=row, column=2, value=f'役員名{i+1}')
                    row += 1
                    ws.cell(row=row, column=2, value=f'ふりがな{i+1}')
                    row += 1
            
            # ◆ VI. 拠点・展開規模
            ws.cell(row=row, column=1, value='◆ VI. 拠点・展開規模')
            ws.cell(row=row, column=2, value='事業所数')
            ws.cell(row=row, column=3, value=company.office_count or '')
            row += 1
            
            ws.cell(row=row, column=2, value='店舗数')
            ws.cell(row=row, column=3, value=company.store_count or '')
            row += 1
            
            # ◆ VII. 拠点・事業所一覧
            for i in range(15):  # 1 to 15
                if i < len(offices):
                    if i == 0:
                        ws.cell(row=row, column=1, value='◆ VII. 拠点・事業所一覧')
                    ws.cell(row=row, column=2, value=f'事業所名{i+1}')
                    ws.cell(row=row, column=3, value=offices[i].name or '')
                    row += 1
                    ws.cell(row=row, column=2, value=f'郵便番号{i+1}')
                    ws.cell(row=row, column=3, value=offices[i].postal_code or '')
                    row += 1
                    ws.cell(row=row, column=2, value=f'住所{i+1}')
                    ws.cell(row=row, column=3, value=offices[i].address or '')
                    row += 1
                    ws.cell(row=row, column=2, value=f'電話番号{i+1}')
                    ws.cell(row=row, column=3, value=offices[i].phone or '')
                    row += 1
                    ws.cell(row=row, column=2, value=f'扱い品目・業務内容{i+1}')
                    ws.cell(row=row, column=3, value=offices[i].business_content or '')
                    row += 1
                else:
                    if i == 0:
                        ws.cell(row=row, column=1, value='◆ VII. 拠点・事業所一覧')
                    ws.cell(row=row, column=2, value=f'事業所名{i+1}')
                    row += 1
                    ws.cell(row=row, column=2, value=f'郵便番号{i+1}')
                    row += 1
                    ws.cell(row=row, column=2, value=f'住所{i+1}')
                    row += 1
                    ws.cell(row=row, column=2, value=f'電話番号{i+1}')
                    row += 1
                    ws.cell(row=row, column=2, value=f'扱い品目・業務内容{i+1}')
                    row += 1
            
            # ◆ VII. URL
            ws.cell(row=row, column=1, value='◆ VII. URL')
            ws.cell(row=row, column=2, value='会社概要ページURL')
            ws.cell(row=row, column=3, value=company.company_overview_url or '')
            row += 1
            
            ws.cell(row=row, column=2, value='拠点・事業所ページURL')
            ws.cell(row=row, column=3, value=company.office_list_url or '')
            row += 1
            
            ws.cell(row=row, column=2, value='組織図ページURL')
            ws.cell(row=row, column=3, value=company.organization_chart_url or '')
            row += 1
            
            ws.cell(row=row, column=2, value='関係会社ページURL')
            ws.cell(row=row, column=3, value=company.related_companies_url or '')
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="companies_detailed.xlsx"'
        return response
        
    except ImportError:
        # Fallback to CSV if openpyxl not available
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="companies_detailed.csv"'
        
        writer = csv.writer(response)
    
    companies = Company.objects.all().order_by('id')
    
    for idx, company in enumerate(companies):
        # Get related executives and offices
        executives = company.executives.all().order_by('order')
        offices = company.offices.all().order_by('order')
        
        # ◆ I. 基本法人情報（識別・概要）
        writer.writerow(['◆ I. 基本法人情報（識別・概要）', '法人番号', company.corporate_number or ''])
        writer.writerow(['', '会社名', company.company_name or ''])
        writer.writerow(['', '会社名かな', company.company_name_kana or ''])
        writer.writerow(['', '英文企業名', company.english_name or ''])
        writer.writerow(['', '代表者名', company.representative_name or ''])
        writer.writerow(['', '代表者かな', company.representative_kana or ''])
        writer.writerow(['', '代表者年齢', company.representative_age or ''])
        writer.writerow(['', '代表者生年月日', company.representative_birth or ''])
        writer.writerow(['', '代表者出身大学', company.representative_university or ''])
        writer.writerow(['', '郵便番号', company.postal_code or ''])
        writer.writerow(['', '住所', company.address or ''])
        writer.writerow(['', '電話番号', company.phone or ''])
        writer.writerow(['', '登記住所', company.registered_address or ''])
        writer.writerow(['', 'FAX番号', company.fax or ''])
        writer.writerow(['', 'URL', company.url or ''])
        writer.writerow(['', '創業', company.founded or ''])
        writer.writerow(['', '設立', company.established or ''])
        writer.writerow(['', '資本金', company.capital or ''])
        writer.writerow(['', '出資金', company.investment or ''])
        writer.writerow(['', '会員数', company.member_count or ''])
        writer.writerow(['', '組合員数', company.union_member_count or ''])
        writer.writerow(['', '上場市場', company.stock_market or ''])
        writer.writerow(['', '証券コード', company.stock_code or ''])
        writer.writerow(['', '決算期', company.fiscal_year_end or ''])
        
        # ◆ II. 経営・財務情報
        writer.writerow(['◆ II. 経営・財務情報', '売上高', company.revenue or ''])
        writer.writerow(['', '純利益', company.net_profit or ''])
        writer.writerow(['', '預金量', company.deposits or ''])
        writer.writerow(['', '従業員数', company.employee_count or ''])
        writer.writerow(['', '平均年齢', company.average_age or ''])
        writer.writerow(['', '平均年収', company.average_salary or ''])
        writer.writerow(['', '役員数', company.executive_count or ''])
        writer.writerow(['', '株主数', company.shareholder_count or ''])
        writer.writerow(['', '取引銀行', company.main_bank or ''])
        
        # ◆ III. 事業・業務内容
        writer.writerow(['◆ III. 事業・業務内容', '業種', company.industry or ''])
        writer.writerow(['', '事業内容', company.business_content or ''])
        writer.writerow(['', '主要事業', company.main_business or ''])
        writer.writerow(['', '事業エリア', company.business_area or ''])
        writer.writerow(['', '系列', company.group_affiliation or ''])
        writer.writerow(['', '販売先', company.sales_destination or ''])
        writer.writerow(['', '仕入先', company.supplier or ''])
        
        # ◆ IV. 役員名簿
        writer.writerow(['◆ IV. 役員名簿', '役職名1', executives[0].position if len(executives) > 0 else ''])
        writer.writerow(['', '役員名1', executives[0].name if len(executives) > 0 else ''])
        writer.writerow(['', 'ふりがな1', executives[0].name_kana if len(executives) > 0 else ''])
        
        for i in range(1, 15):  # 2 to 15
            if i < len(executives):
                writer.writerow(['', f'役職名{i+1}', executives[i].position or ''])
                writer.writerow(['', f'役員名{i+1}', executives[i].name or ''])
                writer.writerow(['', f'ふりがな{i+1}', executives[i].name_kana or ''])
            else:
                writer.writerow(['', f'役職名{i+1}', ''])
                writer.writerow(['', f'役員名{i+1}', ''])
                writer.writerow(['', f'ふりがな{i+1}', ''])
        
        # ◆ VI. 拠点・展開規模
        writer.writerow(['◆ VI. 拠点・展開規模', '事業所数', company.office_count or ''])
        writer.writerow(['', '店舗数', company.store_count or ''])
        
        # ◆ VII. 拠点・事業所一覧
        for i in range(15):  # 1 to 15
            if i < len(offices):
                writer.writerow(['◆ VII. 拠点・事業所一覧' if i == 0 else '', f'事業所名{i+1}', offices[i].name or ''])
                writer.writerow(['', f'郵便番号{i+1}', offices[i].postal_code or ''])
                writer.writerow(['', f'住所{i+1}', offices[i].address or ''])
                writer.writerow(['', f'電話番号{i+1}', offices[i].phone or ''])
                writer.writerow(['', f'扱い品目・業務内容{i+1}', offices[i].business_content or ''])
            else:
                writer.writerow(['◆ VII. 拠点・事業所一覧' if i == 0 else '', f'事業所名{i+1}', ''])
                writer.writerow(['', f'郵便番号{i+1}', ''])
                writer.writerow(['', f'住所{i+1}', ''])
                writer.writerow(['', f'電話番号{i+1}', ''])
                writer.writerow(['', f'扱い品目・業務内容{i+1}', ''])
        
        # ◆ VII. URL
        writer.writerow(['◆ VII. URL', '会社概要ページURL', company.company_overview_url or ''])
        writer.writerow(['', '拠点・事業所ページURL', company.office_list_url or ''])
        writer.writerow(['', '組織図ページURL', company.organization_chart_url or ''])
        writer.writerow(['', '関係会社ページURL', company.related_companies_url or ''])
        
        # Add separator between companies (empty row)
        if idx < len(companies) - 1:  # Don't add separator after last company
            writer.writerow([])
    
    return response

@login_required
def export_single_company_detailed_csv(request, company_id):
    """Export single company detailed CSV file in structured Japanese format"""
    company = get_object_or_404(Company, pk=company_id)
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="company_detail_{company.corporate_number}.csv"'
    
    writer = csv.writer(response)
    
    # Get related executives and offices
    executives = company.executives.all().order_by('order')
    offices = company.offices.all().order_by('order')
    
    # ◆ I. 基本法人情報（識別・概要）
    writer.writerow(['◆ I. 基本法人情報（識別・概要）', '法人番号', company.corporate_number or ''])
    writer.writerow(['', '会社名', company.company_name or ''])
    writer.writerow(['', '会社名かな', company.company_name_kana or ''])
    writer.writerow(['', '英文企業名', company.english_name or ''])
    writer.writerow(['', '代表者名', company.representative_name or ''])
    writer.writerow(['', '代表者かな', company.representative_kana or ''])
    writer.writerow(['', '代表者年齢', company.representative_age or ''])
    writer.writerow(['', '代表者生年月日', company.representative_birth or ''])
    writer.writerow(['', '代表者出身大学', company.representative_university or ''])
    writer.writerow(['', '郵便番号', company.postal_code or ''])
    writer.writerow(['', '住所', company.address or ''])
    writer.writerow(['', '電話番号', company.phone or ''])
    writer.writerow(['', '登記住所', company.registered_address or ''])
    writer.writerow(['', 'FAX番号', company.fax or ''])
    writer.writerow(['', 'URL', company.url or ''])
    writer.writerow(['', '創業', company.founded or ''])
    writer.writerow(['', '設立', company.established or ''])
    writer.writerow(['', '資本金', company.capital or ''])
    writer.writerow(['', '出資金', company.investment or ''])
    writer.writerow(['', '会員数', company.member_count or ''])
    writer.writerow(['', '組合員数', company.union_member_count or ''])
    writer.writerow(['', '上場市場', company.stock_market or ''])
    writer.writerow(['', '証券コード', company.stock_code or ''])
    writer.writerow(['', '決算期', company.fiscal_year_end or ''])
    
    # ◆ II. 経営・財務情報
    writer.writerow(['◆ II. 経営・財務情報', '売上高', company.revenue or ''])
    writer.writerow(['', '純利益', company.net_profit or ''])
    writer.writerow(['', '預金量', company.deposits or ''])
    writer.writerow(['', '従業員数', company.employee_count or ''])
    writer.writerow(['', '平均年齢', company.average_age or ''])
    writer.writerow(['', '平均年収', company.average_salary or ''])
    writer.writerow(['', '役員数', company.executive_count or ''])
    writer.writerow(['', '株主数', company.shareholder_count or ''])
    writer.writerow(['', '取引銀行', company.main_bank or ''])
    
    # ◆ III. 事業・業務内容
    writer.writerow(['◆ III. 事業・業務内容', '業種', company.industry or ''])
    writer.writerow(['', '事業内容', company.business_content or ''])
    writer.writerow(['', '主要事業', company.main_business or ''])
    writer.writerow(['', '事業エリア', company.business_area or ''])
    writer.writerow(['', '系列', company.group_affiliation or ''])
    writer.writerow(['', '販売先', company.sales_destination or ''])
    writer.writerow(['', '仕入先', company.supplier or ''])
    
    # ◆ IV. 役員名簿
    writer.writerow(['◆ IV. 役員名簿', '役職名1', executives[0].position if len(executives) > 0 else ''])
    writer.writerow(['', '役員名1', executives[0].name if len(executives) > 0 else ''])
    writer.writerow(['', 'ふりがな1', executives[0].name_kana if len(executives) > 0 else ''])
    
    for i in range(1, 15):  # 2 to 15
        if i < len(executives):
            writer.writerow(['', f'役職名{i+1}', executives[i].position or ''])
            writer.writerow(['', f'役員名{i+1}', executives[i].name or ''])
            writer.writerow(['', f'ふりがな{i+1}', executives[i].name_kana or ''])
        else:
            writer.writerow(['', f'役職名{i+1}', ''])
            writer.writerow(['', f'役員名{i+1}', ''])
            writer.writerow(['', f'ふりがな{i+1}', ''])
    
    # ◆ VI. 拠点・展開規模
    writer.writerow(['◆ VI. 拠点・展開規模', '事業所数', company.office_count or ''])
    writer.writerow(['', '店舗数', company.store_count or ''])
    
    # ◆ VII. 拠点・事業所一覧
    for i in range(15):  # 1 to 15
        if i < len(offices):
            writer.writerow(['◆ VII. 拠点・事業所一覧' if i == 0 else '', f'事業所名{i+1}', offices[i].name or ''])
            writer.writerow(['', f'郵便番号{i+1}', offices[i].postal_code or ''])
            writer.writerow(['', f'住所{i+1}', offices[i].address or ''])
            writer.writerow(['', f'電話番号{i+1}', offices[i].phone or ''])
            writer.writerow(['', f'扱い品目・業務内容{i+1}', offices[i].business_content or ''])
        else:
            writer.writerow(['◆ VII. 拠点・事業所一覧' if i == 0 else '', f'事業所名{i+1}', ''])
            writer.writerow(['', f'郵便番号{i+1}', ''])
            writer.writerow(['', f'住所{i+1}', ''])
            writer.writerow(['', f'電話番号{i+1}', ''])
            writer.writerow(['', f'扱い品目・業務内容{i+1}', ''])
    
    # ◆ VII. URL
    writer.writerow(['◆ VII. URL', '会社概要ページURL', company.company_overview_url or ''])
    writer.writerow(['', '拠点・事業所ページURL', company.office_list_url or ''])
    writer.writerow(['', '組織図ページURL', company.organization_chart_url or ''])
    writer.writerow(['', '関係会社ページURL', company.related_companies_url or ''])
    
    return response

@login_required
def scraping_history(request):
    """View for scraping execution history"""
    executions = ExecutionHistory.objects.all()
    
    paginator = Paginator(executions.order_by('-started_at'), 20)
    page = paginator.get_page(request.GET.get('page'))
    
    return render(request, 'admin/scraping_history.html', {
        'executions': page,
    })

@login_required
def execution_detail(request, execution_id):
    """View execution details and allow Excel download"""
    execution = get_object_or_404(ExecutionHistory, id=execution_id)
    
    # Get companies processed in this execution (approximate by time range)
    if execution.completed_at:
        companies = Company.objects.filter(
            updated_at__gte=execution.started_at,
            updated_at__lte=execution.completed_at
        ).order_by('-updated_at')
    else:
        companies = Company.objects.filter(
            updated_at__gte=execution.started_at
        ).order_by('-updated_at')[:execution.processed_companies]
    
    return render(request, 'admin/execution_detail.html', {
        'execution': execution,
        'companies': companies[:50],  # Limit display
        'total_companies': companies.count(),
    })

@login_required
def export_execution_data(request, execution_id):
    """Export all data from specific execution"""
    execution = get_object_or_404(ExecutionHistory, id=execution_id)
    
    # Get companies from this execution
    if execution.completed_at:
        companies = Company.objects.filter(
            updated_at__gte=execution.started_at,
            updated_at__lte=execution.completed_at
        ).order_by('-updated_at')
    else:
        companies = Company.objects.filter(
            updated_at__gte=execution.started_at
        ).order_by('-updated_at')[:execution.processed_companies]
    
    try:
        import openpyxl
        from openpyxl.styles import NamedStyle
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        
        # Create text style to prevent scientific notation
        text_style = NamedStyle(name="text_style")
        text_style.number_format = '@'  # Text format
        
        # Companies sheet
        ws1 = wb.active
        ws1.title = "Companies"
        
        headers = ['Corporate Number', 'Company Name', 'Representative', 'Industry', 'Employees', 'Revenue', 'Address', 'Phone', 'Updated']
        for col, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=header)
        
        for row, company in enumerate(companies, 2):
            # Format corporate number as text to prevent scientific notation
            corp_cell = ws1.cell(row=row, column=1, value=company.corporate_number or '')
            corp_cell.number_format = '@'
            
            ws1.cell(row=row, column=2, value=company.company_name or '')
            ws1.cell(row=row, column=3, value=company.representative_name or '')
            ws1.cell(row=row, column=4, value=company.industry or '')
            ws1.cell(row=row, column=5, value=company.employee_count or '')
            ws1.cell(row=row, column=6, value=company.revenue or '')
            ws1.cell(row=row, column=7, value=company.address or '')
            ws1.cell(row=row, column=8, value=company.phone or '')
            ws1.cell(row=row, column=9, value=company.updated_at.strftime('%Y-%m-%d %H:%M:%S') if company.updated_at else '')
        
        # Executives sheet
        ws2 = wb.create_sheet("Executives")
        exec_headers = ['Corporate Number', 'Company Name', 'Position', 'Name', 'Name Kana', 'Order']
        for col, header in enumerate(exec_headers, 1):
            ws2.cell(row=1, column=col, value=header)
        
        exec_row = 2
        for company in companies:
            for executive in company.executives.all():
                corp_cell = ws2.cell(row=exec_row, column=1, value=company.corporate_number)
                corp_cell.number_format = '@'
                
                ws2.cell(row=exec_row, column=2, value=company.company_name)
                ws2.cell(row=exec_row, column=3, value=executive.position)
                ws2.cell(row=exec_row, column=4, value=executive.name)
                ws2.cell(row=exec_row, column=5, value=executive.name_kana)
                ws2.cell(row=exec_row, column=6, value=executive.order)
                exec_row += 1
        
        # Offices sheet
        ws3 = wb.create_sheet("Offices")
        office_headers = ['Corporate Number', 'Company Name', 'Office Name', 'Postal Code', 'Address', 'Phone', 'Business Content', 'Order']
        for col, header in enumerate(office_headers, 1):
            ws3.cell(row=1, column=col, value=header)
        
        office_row = 2
        for company in companies:
            for office in company.offices.all():
                corp_cell = ws3.cell(row=office_row, column=1, value=company.corporate_number)
                corp_cell.number_format = '@'
                
                ws3.cell(row=office_row, column=2, value=company.company_name)
                ws3.cell(row=office_row, column=3, value=office.name)
                ws3.cell(row=office_row, column=4, value=office.postal_code)
                ws3.cell(row=office_row, column=5, value=office.address)
                ws3.cell(row=office_row, column=6, value=office.phone)
                ws3.cell(row=office_row, column=7, value=office.business_content)
                ws3.cell(row=office_row, column=8, value=office.order)
                office_row += 1
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"execution_{execution_id}_{execution.started_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except ImportError:
        # Fallback to CSV with proper encoding
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')  # UTF-8 with BOM
        filename = f"execution_{execution_id}_{execution.started_at.strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['Corporate Number', 'Company Name', 'Representative', 'Industry', 'Employees', 'Revenue', 'Address', 'Updated'])
        
        for company in companies:
            writer.writerow([
                f"'{company.corporate_number}" if company.corporate_number else '',  # Add quote to force text
                company.company_name or '',
                company.representative_name or '',
                company.industry or '',
                company.employee_count or '',
                company.revenue or '',
                company.address or '',
                company.updated_at.strftime('%Y-%m-%d %H:%M:%S') if company.updated_at else ''
            ])
        
        return response

@login_required
@csrf_exempt
def trigger_scraping(request):
    if request.method == 'POST':
        try:
            import json
            
            # Parse configuration from request body
            config = {}
            if request.content_type == 'application/json' and request.body:
                config = json.loads(request.body)
            
            # Create execution record with user configuration
            execution = ExecutionHistory.objects.create(
                spreadsheet_range=config.get('spreadsheet_range', '会社リスト!A3:D'),
                update_google_sheets=config.get('update_google_sheets', True),
                description=config.get('description', ''),
                max_companies=config.get('max_companies', None)
            )
            
            # Start background scraping
            def background_scrape():
                try:
                    from .scraper import CompanyScraper
                    scraper = CompanyScraper()
                    
                    # Check if user provided custom configuration
                    if (execution.spreadsheet_range != '会社リスト!A3:D' or 
                        not execution.update_google_sheets or 
                        execution.max_companies or 
                        execution.description):
                        # Use configured scraping
                        scraper.user_range = execution.spreadsheet_range
                        scraper.user_update_sheets = execution.update_google_sheets
                        scraper.user_max_companies = execution.max_companies
                        result = scraper.scrape_companies_with_config()
                    else:
                        # Use default scraping
                        result = scraper.scrape_companies()
                    
                    execution.status = 'completed'
                    execution.processed_companies = result.get('processed', 0)
                    execution.total_companies = result.get('total', result.get('processed', 0))
                    
                    # Store logs for debugging
                    logs = result.get('logs', [])
                    if logs:
                        execution.error_message = '\n'.join(logs[-50:])  # Last 50 log entries
                    
                    execution.completed_at = timezone.now()
                    execution.save()
                    
                except Exception as e:
                    import traceback
                    error_details = f"Error: {str(e)}\nTraceback: {traceback.format_exc()}"
                    execution.status = 'failed'
                    execution.error_message = error_details
                    execution.completed_at = timezone.now()
                    execution.save()
            
            thread = threading.Thread(target=background_scrape)
            thread.daemon = True
            thread.start()
            
            return JsonResponse({
                'status': 'started', 
                'execution_id': execution.id,
                'message': 'Scraping started with user configuration',
                'config': {
                    'range': execution.spreadsheet_range,
                    'update_sheets': execution.update_google_sheets,
                    'description': execution.description,
                    'max_companies': execution.max_companies
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'POST required'})

@login_required
def api_stats(request):
    return JsonResponse({
        'companies': Company.objects.count(),
        'executives': Executive.objects.count(),
        'offices': Office.objects.count(),
        'recent_changes': ResearchHistory.objects.count(),
    })
